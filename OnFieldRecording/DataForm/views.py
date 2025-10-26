from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime
from io import BytesIO

# PDF and Excel generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Operation, Record, RecordMedia, AuditLog
from .forms import (
    CustomLoginForm, CustomPasswordChangeForm, OperationForm,
    RecordForm, RecordMediaForm, RecordSearchForm
)
from .decorators import staff_required, admin_required, active_operation_required, staff_can_edit_record
from .utils import generate_record_number


# =============================================
# AUTHENTICATION VIEWS
# =============================================

def user_login(request):
    """Login view for all users"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CustomLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            
            # Log the login action
            AuditLog.objects.create(
                user=user,
                action_type='login',
                target_type='user',
                target_id=user.id,
                details={'username': user.username},
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Welcome back, {user.username}!')
            
            # Redirect based on role
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            
            # Admin users go to dashboard
            if hasattr(user, 'profile') and user.profile.role == 'admin':
                return redirect('dashboard')
            
            # Staff users go directly to record creation
            # (active_operation_required decorator will handle checking for active operation)
            return redirect('record_create')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = CustomLoginForm()
    
    return render(request, 'registration/login.html', {'form': form})


@login_required
def user_logout(request):
    """Logout view"""
    # Log the logout action
    AuditLog.objects.create(
        user=request.user,
        action_type='logout',
        target_type='user',
        target_id=request.user.id,
        details={'username': request.user.username},
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


@login_required
def change_password(request):
    """Password change view"""
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Keep user logged in
            
            # Log password change
            AuditLog.objects.create(
                user=request.user,
                action_type='password_change',
                target_type='user',
                target_id=request.user.id,
                details={'username': request.user.username},
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, 'Your password was successfully updated!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomPasswordChangeForm(request.user)
    
    return render(request, 'registration/password_change.html', {'form': form})


# =============================================
# DASHBOARD & HOME
# =============================================

@login_required
def dashboard(request):
    """Main dashboard - shows different content based on role"""
    user = request.user
    context = {
        'user': user,
        'is_admin': hasattr(user, 'profile') and user.profile.role == 'admin'
    }
    
    # Note: active_operation is now provided by context processor
    
    if context['is_admin']:
        # Admin dashboard - show all operations and stats
        operations = Operation.objects.filter(is_deleted=False).order_by('-created_at')[:10]
        
        # Overall stats
        total_operations = Operation.objects.filter(is_deleted=False).count()
        total_records = Record.objects.filter(is_deleted=False).count()
        
        # Recent records
        recent_records = Record.objects.filter(is_deleted=False).order_by('-created_at')[:10]
        
        # Anomaly distribution
        anomaly_stats = Record.objects.filter(is_deleted=False).values('type_of_anomaly').annotate(
            count=Count('id')
        ).order_by('-count')
        
        context.update({
            'operations': operations,
            'total_operations': total_operations,
            'total_records': total_records,
            'recent_records': recent_records,
            'anomaly_stats': anomaly_stats,
        })
        
        return render(request, 'dataform/admin_dashboard.html', context)
    else:
        # Staff dashboard - show their records
        user_records = Record.objects.filter(
            created_by=user,
            is_deleted=False
        ).order_by('-created_at')[:20]
        
        user_stats = {
            'total_records': Record.objects.filter(created_by=user, is_deleted=False).count(),
            'draft_count': Record.objects.filter(created_by=user, status='draft', is_deleted=False).count(),
            'submitted_count': Record.objects.filter(created_by=user, status='submitted', is_deleted=False).count(),
        }
        
        context.update({
            'user_records': user_records,
            'user_stats': user_stats,
        })
        
        return render(request, 'dataform/staff_dashboard.html', context)


# =============================================
# OPERATION VIEWS (Admin Only)
# =============================================

@admin_required
def operation_list(request):
    """List all operations"""
    operations = Operation.objects.filter(is_deleted=False).order_by('-created_at')
    
    # Add record counts
    for op in operations:
        op.record_count = Record.objects.filter(operation=op, is_deleted=False).count()
    
    context = {
        'operations': operations,
    }
    return render(request, 'dataform/operation_list.html', context)


@admin_required
def operation_create(request):
    """Create a new operation"""
    if request.method == 'POST':
        form = OperationForm(request.POST)
        if form.is_valid():
            operation = form.save(commit=False)
            operation.created_by = request.user
            
            try:
                operation.save()
                messages.success(request, f'Operation "{operation.name}" created successfully!')
                return redirect('operation_detail', pk=operation.pk)
            except Exception as e:
                messages.error(request, f'Error creating operation: {str(e)}')
    else:
        form = OperationForm()
    
    return render(request, 'dataform/operation_form.html', {'form': form, 'action': 'Create'})


@admin_required
def operation_detail(request, pk):
    """View operation details"""
    operation = get_object_or_404(Operation, pk=pk)
    
    # Get records for this operation
    records = Record.objects.filter(operation=operation, is_deleted=False).order_by('-created_at')[:50]
    
    # Stats
    stats = {
        'total_records': Record.objects.filter(operation=operation, is_deleted=False).count(),
        'draft': Record.objects.filter(operation=operation, status='draft', is_deleted=False).count(),
        'submitted': Record.objects.filter(operation=operation, status='submitted', is_deleted=False).count(),
        'verified': Record.objects.filter(operation=operation, status='verified', is_deleted=False).count(),
        'with_anomaly': Record.objects.filter(operation=operation, is_deleted=False).exclude(type_of_anomaly='none').count(),
    }
    
    # Anomaly distribution
    from django.db.models import Count
    anomaly_stats = Record.objects.filter(
        operation=operation, 
        is_deleted=False
    ).exclude(
        type_of_anomaly='none'
    ).values('type_of_anomaly').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'operation': operation,
        'records': records,
        'stats': stats,
        'anomaly_stats': anomaly_stats,
    }
    return render(request, 'dataform/operation_detail.html', context)


@admin_required
def operation_activate(request, pk):
    """Activate an operation"""
    operation = get_object_or_404(Operation, pk=pk)
    
    try:
        operation.reopen_operation()
        messages.success(request, f'Operation "{operation.name}" has been activated.')
    except Exception as e:
        messages.error(request, f'Error activating operation: {str(e)}')
    
    return redirect('operation_detail', pk=pk)


@admin_required
def operation_close(request, pk):
    """Close an operation"""
    operation = get_object_or_404(Operation, pk=pk)
    
    operation.close_operation(request.user)
    messages.success(request, f'Operation "{operation.name}" has been closed.')
    
    return redirect('operation_detail', pk=pk)


@admin_required
def operation_delete(request, pk):
    """Delete an operation (admin only) with audit logging"""
    operation = get_object_or_404(Operation, pk=pk)
    
    if request.method == 'POST':
        from ..OnFieldRecording.DataForm.models import DeletionLog
        
        operation_name = operation.name
        operation_id = operation.pk
        
        # Get deletion reason from form
        deletion_reason = request.POST.get('deletion_reason', '').strip()
        
        # Count related records
        record_count = Record.objects.filter(operation=operation, is_deleted=False).count()
        
        # Collect metadata before deletion
        metadata = {
            'operation_name': operation_name,
            'record_count': record_count,
            'was_active': operation.is_active,
            'created_by': operation.created_by.username if operation.created_by else None,
            'created_at': operation.created_at.isoformat(),
            'next_record_seq': operation.next_record_seq,
        }
        
        # Log the deletion BEFORE actually deleting
        DeletionLog.objects.create(
            deleted_by=request.user,
            item_type='operation',
            item_id=operation_id,
            item_name=operation_name,
            deletion_reason=deletion_reason or 'No reason provided',
            metadata=metadata
        )
        
        # Delete the operation (this will cascade delete related records)
        operation.delete()
        
        messages.success(request, f'Operation "{operation_name}" and its {record_count} record(s) have been deleted successfully.')
        return redirect('operation_list')
    
    return redirect('operation_detail', pk=pk)


@admin_required
def operation_export_pdf(request, pk):
    """Export operation details and records to PDF"""
    operation = get_object_or_404(Operation, pk=pk)
    records = Record.objects.filter(operation=operation, is_deleted=False).select_related(
        'created_by', 'operation'
    ).order_by('record_number')
    
    # Get statistics
    total_records = records.count()
    stats = {
        'draft': records.filter(status='draft').count(),
        'submitted': records.filter(status='submitted').count(),
        'verified': records.filter(status='verified').count(),
    }
    
    # Anomaly stats
    anomaly_stats = records.exclude(type_of_anomaly='none').values('type_of_anomaly').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Create filename from operation name (sanitize for filesystem)
    safe_operation_name = "".join(c for c in operation.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    safe_operation_name = safe_operation_name.replace(' ', '_')
    filename = f"{safe_operation_name}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create the PDF object using ReportLab with LANDSCAPE orientation
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    
    # Container for PDF elements
    elements = []
    
    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#374151'),
        spaceAfter=12,
        spaceBefore=20
    )
    normal_style = styles['Normal']
    
    # Title
    title = Paragraph(f"<b>Operation Report</b>", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Operation Details Table
    operation_data = [
        ['Operation Name:', operation.name],
        ['Description:', operation.description or 'N/A'],
        ['Status:', 'Active' if operation.is_active else 'Inactive'],
        ['Created By:', operation.created_by.username if operation.created_by else 'N/A'],
        ['Start Date:', operation.start_at.strftime('%Y-%m-%d %H:%M') if operation.start_at else 'N/A'],
        ['End Date:', operation.end_at.strftime('%Y-%m-%d %H:%M') if operation.end_at else 'N/A'],
        ['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M')],
    ]
    
    operation_table = Table(operation_data, colWidths=[2*inch, 4.5*inch])
    operation_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
    ]))
    
    elements.append(operation_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistics Section
    elements.append(Paragraph("<b>Statistics</b>", heading_style))
    
    stats_data = [
        ['Total Records', 'Draft', 'Submitted', 'Verified', 'Anomalies'],
        [str(total_records), str(stats['draft']), str(stats['submitted']), str(stats['verified']), str(len(anomaly_stats))]
    ]
    
    stats_table = Table(stats_data, colWidths=[1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Anomaly Breakdown (if any)
    if anomaly_stats:
        elements.append(Paragraph("<b>Anomaly Breakdown</b>", heading_style))
        
        anomaly_data = [['Anomaly Type', 'Count']]
        for anomaly in anomaly_stats:
            anomaly_type = anomaly['type_of_anomaly'].replace('_', ' ').title()
            anomaly_data.append([anomaly_type, str(anomaly['count'])])
        
        anomaly_table = Table(anomaly_data, colWidths=[4*inch, 2.5*inch])
        anomaly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
        ]))
        
        elements.append(anomaly_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Records Table (all records)
    if total_records > 0:
        elements.append(Paragraph("<b>Records</b>", heading_style))
        
        # Table headers - only specified fields
        records_data = [[
            'Job #', 
            'Customer Name', 
            'Contact', 
            'GPS Address', 
            'Account #', 
            'Meter #', 
            'Balance', 
            'Reading', 
            'Anomaly', 
            'Remarks'
        ]]
        
        # Add all records with specified fields
        for idx, record in enumerate(records, 1):
            anomaly_display = record.type_of_anomaly.replace('_', ' ').title() if record.type_of_anomaly != 'none' else 'None'
            records_data.append([
                record.record_number,
                record.customer_name[:30] or 'N/A',  # Increased from 25
                record.customer_contact or 'N/A',
                (record.gps_address[:40] + '...') if len(record.gps_address) > 40 else (record.gps_address or 'N/A'),  # Increased from 30
                record.account_number or 'N/A',
                record.meter_number or 'N/A',
                f"{record.todays_balance:,.2f}",
                f"{record.meter_reading:,.2f}",
                anomaly_display,
                (record.remarks[:25] + '...') if len(record.remarks) > 25 else (record.remarks or '-')  # Increased from 20
            ])
        
        # Adjusted column widths for landscape layout (total ~10 inches)
        records_table = Table(records_data, colWidths=[
            0.95*inch,  # Job #
            1.35*inch,  # Customer Name
            1.0*inch,   # Contact
            1.5*inch,   # GPS Address
            0.95*inch,  # Account #
            0.95*inch,  # Meter #
            0.75*inch,  # Balance
            0.75*inch,  # Reading
            1.0*inch,   # Anomaly
            0.8*inch    # Remarks
        ])
        records_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ]))
        
        elements.append(records_table)
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    # Log export action
    AuditLog.objects.create(
        user=request.user,
        action_type='export',
        target_type='operation',
        target_id=operation.pk,
        details={'format': 'pdf', 'record_count': total_records},
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return response


@admin_required
def operation_export_xlsx(request, pk):
    """Export operation details and records to Excel (XLSX)"""
    operation = get_object_or_404(Operation, pk=pk)
    records = Record.objects.filter(operation=operation, is_deleted=False).select_related(
        'created_by', 'operation'
    ).order_by('record_number')
    
    # Get statistics
    total_records = records.count()
    stats = {
        'draft': records.filter(status='draft').count(),
        'submitted': records.filter(status='submitted').count(),
        'verified': records.filter(status='verified').count(),
    }
    
    # Anomaly stats
    anomaly_stats = records.exclude(type_of_anomaly='none').values('type_of_anomaly').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Create workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # ========== SHEET 1: Summary ==========
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = "Operation Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    ws_summary.merge_cells('A1:B1')
    ws_summary.row_dimensions[1].height = 25
    
    # Operation Details
    ws_summary['A3'] = "Operation Details"
    ws_summary['A3'].font = Font(bold=True, size=12)
    ws_summary.merge_cells('A3:B3')
    
    details = [
        ['Operation Name:', operation.name],
        ['Description:', operation.description or 'N/A'],
        ['Status:', 'Active' if operation.is_active else 'Inactive'],
        ['Created By:', operation.created_by.username if operation.created_by else 'N/A'],
        ['Start Date:', operation.start_at.strftime('%Y-%m-%d %H:%M') if operation.start_at else 'N/A'],
        ['End Date:', operation.end_at.strftime('%Y-%m-%d %H:%M') if operation.end_at else 'N/A'],
        ['Total Records:', str(total_records)],
        ['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    row = 4
    for label, value in details:
        ws_summary[f'A{row}'] = label
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'A{row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].border = border
        ws_summary[f'B{row}'].border = border
        row += 1
    
    # Statistics
    row += 1
    ws_summary[f'A{row}'] = "Record Statistics"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    ws_summary.merge_cells(f'A{row}:E{row}')
    
    row += 1
    stats_headers = ['Total Records', 'Draft', 'Submitted', 'Verified', 'Anomalies']
    for col, header in enumerate(stats_headers, 1):
        cell = ws_summary.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row += 1
    stats_values = [total_records, stats['draft'], stats['submitted'], stats['verified'], len(anomaly_stats)]
    for col, value in enumerate(stats_values, 1):
        cell = ws_summary.cell(row, col, value)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Anomaly Breakdown
    if anomaly_stats:
        row += 2
        ws_summary[f'A{row}'] = "Anomaly Breakdown"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        ws_summary.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws_summary[f'A{row}'] = "Anomaly Type"
        ws_summary[f'B{row}'] = "Count"
        for col in range(1, 3):
            cell = ws_summary.cell(row, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        for anomaly in anomaly_stats:
            row += 1
            anomaly_type = anomaly['type_of_anomaly'].replace('_', ' ').title()
            ws_summary[f'A{row}'] = anomaly_type
            ws_summary[f'B{row}'] = anomaly['count']
            ws_summary[f'A{row}'].border = border
            ws_summary[f'B{row}'].border = border
            ws_summary[f'B{row}'].alignment = Alignment(horizontal="center")
    
    # Set column widths for summary sheet
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 40
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    
    # ========== SHEET 2: Records Data ==========
    ws_records = wb.create_sheet(title="Records")
    
    # Headers - only specified fields
    headers = [
        'Job Number', 
        'Customer Name', 
        'Customer Contact', 
        'GPS Address', 
        'Account Number', 
        'Meter Number',
        'Today\'s Balance', 
        'Meter Reading', 
        'Type of Anomaly', 
        'Remarks'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_records.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows with specified fields only
    for idx, record in enumerate(records, 1):
        row_data = [
            record.record_number,
            record.customer_name or '',
            record.customer_contact or '',
            record.gps_address or '',
            record.account_number or '',
            record.meter_number or '',
            float(record.todays_balance) if record.todays_balance else 0,
            float(record.meter_reading) if record.meter_reading else 0,
            record.type_of_anomaly.replace('_', ' ').title() if record.type_of_anomaly != 'none' else 'None',
            record.remarks or '',
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_records.cell(idx + 1, col, value)
            cell.alignment = cell_alignment
            cell.border = border
            
            # Format currency columns
            if col in [7, 8]:  # Balance and Reading columns
                cell.number_format = '#,##0.00'
    
    # Set column widths for records sheet
    column_widths = {
        'A': 18,  # Job Number
        'B': 25,  # Customer Name
        'C': 18,  # Customer Contact
        'D': 35,  # GPS Address
        'E': 18,  # Account Number
        'F': 18,  # Meter Number
        'G': 15,  # Today's Balance
        'H': 15,  # Meter Reading
        'I': 20,  # Type of Anomaly
        'J': 30,  # Remarks
    }
    
    for col_letter, width in column_widths.items():
        ws_records.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws_records.freeze_panes = 'A2'
    
    # Create sanitized filename with operation name
    safe_operation_name = "".join(c for c in operation.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    safe_operation_name = safe_operation_name.replace(' ', '_')
    timestamp = timezone.now().strftime('%Y%m%d_%H%M')
    filename = f"{safe_operation_name}_{timestamp}.xlsx"
    
    # Create the HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    # Log export action
    AuditLog.objects.create(
        user=request.user,
        action_type='export',
        target_type='operation',
        target_id=operation.pk,
        details={'format': 'xlsx', 'record_count': total_records},
        ip_address=request.META.get('REMOTE_ADDR')
    )
    
    return response


# =============================================
# RECORD VIEWS
# =============================================

@staff_required
@active_operation_required
@staff_required
@active_operation_required
def record_create(request):
    """Create a new record (staff view)"""
    active_operation = request.active_operation
    
    if request.method == 'POST':
        form = RecordForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            record.operation = active_operation
            record.created_by = request.user
            
            # Generate record number
            record.record_number = generate_record_number(active_operation)
            
            record.save()
            
            # Handle photo uploads
            photos = request.FILES.getlist('photos')
            for photo in photos:
                RecordMedia.objects.create(
                    record=record,
                    image=photo,
                    uploaded_by=request.user
                )
            
            messages.success(request, f'Record {record.record_number} created successfully!')
            
            # Check if user wants to create another
            if 'save_and_new' in request.POST:
                return redirect('record_create')
            else:
                return redirect('record_detail', pk=record.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RecordForm(initial={'status': 'draft'})
    
    context = {
        'form': form,
        'active_operation': active_operation,
        'action': 'Create'
    }
    return render(request, 'dataform/record_form.html', context)


@staff_required
def record_list(request):
    """List records with filtering"""
    records = Record.objects.filter(is_deleted=False).select_related('operation', 'created_by')
    
    # Apply filters
    search_form = RecordSearchForm(request.GET)
    
    if search_form.is_valid():
        search = search_form.cleaned_data.get('search')
        if search:
            records = records.filter(
                Q(record_number__icontains=search) |
                Q(customer_name__icontains=search) |
                Q(account_number__icontains=search) |
                Q(meter_number__icontains=search) |
                Q(customer_contact__icontains=search)
            )
        
        operation = search_form.cleaned_data.get('operation')
        if operation:
            records = records.filter(operation=operation)
        
        status = search_form.cleaned_data.get('status')
        if status:
            records = records.filter(status=status)
        
        anomaly = search_form.cleaned_data.get('anomaly')
        if anomaly:
            records = records.filter(type_of_anomaly=anomaly)
        
        date_from = search_form.cleaned_data.get('date_from')
        if date_from:
            records = records.filter(created_at__date__gte=date_from)
        
        date_to = search_form.cleaned_data.get('date_to')
        if date_to:
            records = records.filter(created_at__date__lte=date_to)
    
    # Order by latest first
    records = records.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(records, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'total_count': records.count(),
    }
    return render(request, 'dataform/record_list.html', context)


@staff_required
def record_detail(request, pk):
    """View record details"""
    record = get_object_or_404(Record, pk=pk, is_deleted=False)
    
    # Check permissions
    is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'admin'
    can_edit = is_admin or (record.created_by == request.user and record.operation.is_active)
    
    context = {
        'record': record,
        'can_edit': can_edit,
        'is_admin': is_admin,
    }
    return render(request, 'dataform/record_detail.html', context)


@staff_required
@staff_can_edit_record
def record_update(request, pk):
    """Update a record"""
    record = get_object_or_404(Record, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        form = RecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            
            # Handle new photo uploads
            photos = request.FILES.getlist('photos')
            for photo in photos:
                RecordMedia.objects.create(
                    record=record,
                    image=photo,
                    uploaded_by=request.user
                )
            
            messages.success(request, f'Record {record.record_number} updated successfully!')
            return redirect('record_detail', pk=pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RecordForm(instance=record)
    
    context = {
        'form': form,
        'record': record,
        'action': 'Update'
    }
    return render(request, 'dataform/record_form.html', context)


@login_required
def record_delete(request, pk):
    """Delete a record (admin or record owner only) with audit logging"""
    record = get_object_or_404(Record, pk=pk, is_deleted=False)
    
    # Permission check: must be admin or record creator
    if not (request.user.profile.role == 'admin' or record.created_by == request.user):
        messages.error(request, 'You do not have permission to delete this record.')
        return redirect('record_detail', pk=pk)
    
    if request.method == 'POST':
        from .models import DeletionLog
        
        record_number = record.record_number
        record_id = record.pk
        operation_name = record.operation.name if record.operation else 'None'
        
        # Get deletion reason from form
        deletion_reason = request.POST.get('deletion_reason', '').strip()
        
        # Count related media
        media_count = RecordMedia.objects.filter(record=record).count()
        
        # Collect metadata before deletion
        metadata = {
            'record_number': record_number,
            'operation': operation_name,
            'customer_name': record.customer_name,
            'account_number': record.account_number,
            'meter_number': record.meter_number,
            'status': record.status,
            'type_of_anomaly': record.type_of_anomaly,
            'media_count': media_count,
            'created_by': record.created_by.username if record.created_by else None,
            'created_at': record.created_at.isoformat(),
        }
        
        # Log the deletion BEFORE actually deleting
        DeletionLog.objects.create(
            deleted_by=request.user,
            item_type='record',
            item_id=record_id,
            item_name=record_number,
            deletion_reason=deletion_reason or 'No reason provided',
            metadata=metadata
        )
        
        # Delete the record (this will cascade delete related media)
        record.delete()
        
        messages.success(request, f'Record {record_number} has been deleted successfully.')
        return redirect('record_list')
    
    return redirect('record_detail', pk=pk)


# =============================================
# API ENDPOINTS (for AJAX)
# =============================================

@login_required
def get_active_operation(request):
    """API endpoint to get current active operation"""
    active_op = Operation.objects.filter(is_active=True, is_deleted=False).first()
    
    if active_op:
        return JsonResponse({
            'id': active_op.id,
            'name': active_op.name,
            'description': active_op.description,
        })
    else:
        return JsonResponse({'error': 'No active operation'}, status=404)


# =============================================
# SEARCH FUNCTIONALITY
# =============================================

@admin_required
def operation_search(request, pk):
    """Search records within a specific operation (Admin only)"""
    operation = get_object_or_404(Operation, pk=pk, is_deleted=False)
    query = request.GET.get('q', '').strip()
    
    # Start with records in this operation
    records = Record.objects.filter(operation=operation, is_deleted=False).select_related(
        'created_by', 'operation'
    )
    
    # Apply search filter if query exists
    if query:
        records = records.filter(
            Q(customer_name__icontains=query) |
            Q(customer_contact__icontains=query) |
            Q(account_number__icontains=query) |
            Q(meter_number__icontains=query) |
            Q(gps_address__icontains=query) |
            Q(record_number__icontains=query) |
            Q(remarks__icontains=query)
        )
    
    # Order by most recent first
    records = records.order_by('-created_at')
    
    # Paginate results
    paginator = Paginator(records, 50)  # 50 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'operation': operation,
        'records': page_obj,
        'query': query,
        'total_results': records.count(),
        'search_type': 'operation',
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    }
    
    return render(request, 'dataform/search_results.html', context)


@admin_required
def system_search(request):
    """Search all records across all operations (Admin only)"""
    query = request.GET.get('q', '').strip()
    
    # Start with all non-deleted records
    records = Record.objects.filter(is_deleted=False).select_related(
        'created_by', 'operation'
    )
    
    # Apply search filter if query exists
    if query:
        records = records.filter(
            Q(customer_name__icontains=query) |
            Q(customer_contact__icontains=query) |
            Q(account_number__icontains=query) |
            Q(meter_number__icontains=query) |
            Q(gps_address__icontains=query) |
            Q(record_number__icontains=query) |
            Q(remarks__icontains=query) |
            Q(operation__name__icontains=query)  # Also search by operation name
        )
    
    # Order by most recent first
    records = records.order_by('-created_at')
    
    # Paginate results
    paginator = Paginator(records, 50)  # 50 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'records': page_obj,
        'query': query,
        'total_results': records.count(),
        'search_type': 'system',
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    }
    
    return render(request, 'dataform/search_results.html', context)
